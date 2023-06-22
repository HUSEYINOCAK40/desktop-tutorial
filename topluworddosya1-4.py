import tkinter as tk
from tkinter import LabelFrame, filedialog
import openpyxl
from docx import Document
from PIL import ImageTk, Image
from docx2pdf import convert
import os
import tkinter.messagebox as messagebox
from tkinter import ttk
import csv
import xlsxwriter
from tkinter import Tk, Frame, Canvas, Scrollbar, Label
from tkinter.font import Font

def scroll_y(event):
    canvas.yview_scroll(-1 * int(event.delta / 120), "units")

def scroll_x(event):
    canvas.xview_scroll(-1 * int(event.delta / 120), "units")

def excel_dosyasini_sec():
    dosya_yolu = filedialog.askopenfilename(filetypes=[("Excel Dosyası", "*.xlsx")])
    excel_dosya_entry.delete(0, tk.END)
    excel_dosya_entry.insert(0, "Seçilen Excel Dosyası: " + dosya_yolu)

def sablon_dosyasini_sec():
    dosya_yolu = filedialog.askopenfilename(filetypes=[("Word Şablonu", "*.docx")])
    sablon_dosya_entry.delete(0, tk.END)
    sablon_dosya_entry.insert(0, dosya_yolu)

def klasor_dosyasini_sec():
    dosya_yolu = filedialog.askdirectory()
    kayit_klasoru_entry.delete(0, tk.END)
    kayit_klasoru_entry.insert(0, dosya_yolu) 

def toplu_pdf_dosyalari_olustur():
    pdf_kayit_klasoru = kayit_klasoru_entry.get()  # Giriş kutusundan kayıt klasörünün yolunu alın
    os.makedirs(pdf_kayit_klasoru, exist_ok=True)  # Kayıt klasörünü oluşturun (varsa zaten geçerli olanı kullanın)
    sayac=0
    for dosya in os.listdir(pdf_kayit_klasoru):
        if dosya.endswith(".docx"):  # Sadece .docx uzantılı dosyalara bakın
            dosya_yolu = os.path.join(pdf_kayit_klasoru, dosya)  # Dosya yolunu oluşturun
            pdf_adi = dosya.replace(".docx", ".pdf")  # Dosya adını .pdf ile değiştirin
            hedef_yolu = os.path.join(pdf_kayit_klasoru, pdf_adi)  # Hedef yolunu oluşturun
            sayac=sayac+1
            try:
                convert(dosya_yolu, hedef_yolu)  # Dosyayı PDF'ye dönüştürün
            except Exception as e:
                messagebox.showerror("Hata", f"{dosya} dönüştürülürken bir hata oluştu:\n{str(e)}")  # Hata durumunda bir hata iletişim kutusu gösterin
                break
    messagebox.showinfo("Bilgi", f"{sayac} dosya PDF'e dönüştürüldü.")  # İşlem tamamlandığında bir bilgi iletişim kutusu gösterin

def csv_to_excel():
    csv_folder = kayit_klasoru_entry.get()
    os.makedirs(csv_folder, exist_ok=True)  # Kayıt klasörünü oluşturun (varsa zaten geçerli olanı kullanın)
    sayac=0
    # CSV klasöründeki dosyaları gezin
    for file in os.listdir(csv_folder):
        if file.endswith(".csv"):
            csv_file = os.path.join(csv_folder, file)  # CSV dosyasının tam yolunu oluşturun
            excel_file = os.path.join(csv_folder, file.replace(".csv", ".xlsx"))  # Excel dosyasının tam yolunu oluşturun
           
            workbook = xlsxwriter.Workbook(excel_file)
            worksheet = workbook.add_worksheet()
            sayac=sayac+1
            try:
                with open(csv_file, 'r', encoding='utf-8') as csvfile:
                    csvreader = csv.reader(csvfile)
                    for row_num, row_data in enumerate(csvreader):
                        worksheet.write_row(row_num, 0, row_data)
                workbook.close()
            except Exception as e:
                messagebox.showerror("Hata", f"{file} dönüştürülürken bir hata oluştu:\n{str(e)}")  # Hata durumunda bir hata iletişim kutusu gösterin
                break
    messagebox.showinfo("Bilgi", f"{sayac} dosya Excel'e dönüştürüldü.")  # İşlem tamamlandığında bir bilgi iletişim kutusu gösterin

def toplu_word_dosyalari_olustur():
    excel_dosya = excel_dosya_entry.get().split(": ")[1]
    word_sablonu = sablon_dosya_entry.get()
    kayit_klasoru = kayit_klasoru_entry.get()
    sayaca=0
    workbook = openpyxl.load_workbook(excel_dosya)
    sheet = workbook.active

    for column in sheet.iter_cols():
        
        veri1 = column[0].value
        veri2 = column[1].value
        veri3 = column[2].value
        veri4 = column[3].value
        veri5 = column[4].value
        veri6 = column[5].value
        veri7 = column[6].value
        veri8 = column[7].value
        veri9 = column[8].value
        veria0 = column[9].value
        veria1= column[10].value
        veria2 = column[11].value
        veria3 = column[12].value
        veria4 = column[13].value
        veria5 = column[14].value
        veria6 = column[15].value
        veria7 = column[16].value
        veria8 = column[17].value
        veria9 = column[18].value
        verib0 = column[19].value
        verib1 = column[20].value
        verib2 = column[21].value
        verib3 = column[22].value
        verib4 = column[23].value
        verib5 = column[24].value
        verib6 = column[25].value
        verib7 = column[26].value
        verib8 = column[27].value
        verib9 = column[28].value

        word_belgesi = Document(word_sablonu)
        # Word belgesinde "veri1" ve "veri2" yerine gerçek verileri yerleştirme işlemleri
        for paragraph in word_belgesi.paragraphs:
            if "veri1" in paragraph.text:
                paragraph.text = paragraph.text.replace("veri1", str(veri1))
            if "veri2" in paragraph.text:
                paragraph.text = paragraph.text.replace("veri2", str(veri2))
            if "veri3" in paragraph.text:
                paragraph.text = paragraph.text.replace("veri3", str(veri3))
            if "veri4" in paragraph.text:
                paragraph.text = paragraph.text.replace("veri4", str(veri4))
            if "veri5" in paragraph.text:
                paragraph.text = paragraph.text.replace("veri5", str(veri5))
            if "veri6" in paragraph.text:
                paragraph.text = paragraph.text.replace("veri6", str(veri6))     
            if "veri7" in paragraph.text:
                paragraph.text = paragraph.text.replace("veri7", str(veri7))
            if "veri8" in paragraph.text:
                paragraph.text = paragraph.text.replace("veri8", str(veri8))
            if "veri9" in paragraph.text:
                paragraph.text = paragraph.text.replace("veri9", str(veri9))
            if "veria0" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria0", str(veria0))
            if "veria1" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria1", str(veria1))            
            if "veria2" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria2", str(veria2))     
            if "veria3" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria3", str(veria3))
            if "veria4" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria4", str(veria4))
            if "veria5" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria5", str(veria5))
            if "veria6" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria6", str(veria6))
            if "veria7" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria7", str(veria7))
            if "veria8" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria8", str(veria8))  
            if "veria9" in paragraph.text:
                paragraph.text = paragraph.text.replace("veria9", str(veria9))      
            if "verib0" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib0", str(verib0))
            if "verib1" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib1", str(verib1))            
            if "verib2" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib2", str(verib2))     
            if "verib3" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib3", str(verib3))
            if "verib4" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib4", str(verib4))
            if "verib5" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib5", str(verib5))
            if "verib6" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib6", str(verib6))
            if "verib7" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib7", str(verib7))
            if "verib8" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib8", str(verib8))  
            if "verib9" in paragraph.text:
                paragraph.text = paragraph.text.replace("verib9", str(verib9))       

    
       # if veri1 != "":
        word_belgesi.save(os.path.join(kayit_klasoru, f"{veri1}.docx"))
        sayaca = sayaca + 1
    messagebox.showinfo("Bilgi", f"{sayaca} Word dosyası oluşturuldu.") 

def toggle_label_frame_visibility():
    if LabelFrame.winfo_viewable():
        ttk.Labelframe.pack_forget()
    else:
        ttk.Labelframe.pack()

# Arayüzü oluşturma
root = tk.Tk()
root.title('ExceltoWord')
root.iconbitmap("icon.ico") 
root.geometry("700x450")

#-------------------------------------------------------

# Dikey scrollbar'ı kontrol etmek için
root.bind("<MouseWheel>", scroll_y)

# Yatay scrollbar'ı kontrol etmek için
root.bind("<Shift-MouseWheel>", scroll_x)

# Dosyadan metin okumak için
file_path = "notlar.txt"
lines = []

# Dosyayı satır satır okuyarak lines listesine ekler
with open(file_path, "r", encoding="utf-8") as file:
    lines = file.readlines()

line_count = len(lines)

# Canvas ve ana frame oluşturma
main_frame = Frame(root)
main_frame.grid(row=10, column=1)

canvas = Canvas(main_frame)
canvas.grid(sticky="news")

# Pencere içine frame oluşturma
inner_frame = Frame(canvas, width=3, height=3)
canvas.create_window((0, 0), window=inner_frame, anchor="nw")

liste = ttk.Treeview(inner_frame, height=3)
liste.grid(sticky="w")

# Frame için scrollbar oluşturma
sb1 = Scrollbar(canvas, orient="vertical", command=canvas.yview)
sb1.place(x=365, y=0, height=135)


sb2 = Scrollbar(canvas, orient="horizontal", command=canvas.xview)
sb2.place(x=0, y=120, width=365)

canvas.configure(yscrollcommand=sb1.set, xscrollcommand=sb2.set)
canvas.grid(row=0, column=1, sticky="news")

sb1.config(command=canvas.yview)
sb2.config(command=canvas.xview)

# Metin dosyasındaki satırları tek tek frame içine bastırma
for i in range(line_count):
    text_label = ttk.Label(liste, text=lines[i], anchor="w")
    text_label.grid(row=i, column=0, sticky="w")
    text_label.configure(background="white", foreground="black", font=("Arial", 9))

# Resim oluşturma
image = Image.open("resim.png")
image = image.resize((300, 300))
tk_image = ImageTk.PhotoImage(image)
image_label = Label(liste, image=tk_image)
image_label.grid(sticky="w")

# Canvas'ın scrollregion'ını güncelleme
inner_frame.update_idletasks()
canvas.configure(scrollregion=canvas.bbox("all"), width=375, height=130)
canvas.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))

#-----------------------------------------------------------

excel_dosya_label = tk.Label(root, text="Excel Dosyası:")
excel_dosya_label.grid(row=0, column=0)

excel_dosya_entry = tk.Entry(root, width=50)
excel_dosya_entry.grid(row=0, column=1)

excel_dosya_sec_button = tk.Button(root, text="Dosya Seç", command=excel_dosyasini_sec)
excel_dosya_sec_button.grid(row=0, column=2)

nota9_label = tk.Label(root, text="          ")
nota9_label.grid(row=0, column=3)

# Word şablonu seçme
sablon_dosya_label = tk.Label(root, text="\n Word Şablonu:")
sablon_dosya_label.grid(row=1, column=0)

sablon_dosya_entry = tk.Entry(root, width=50)
sablon_dosya_entry.grid(row=1, column=1)

sablon_dosya_sec_button = tk.Button(root, text="Dosya Seç", command=sablon_dosyasini_sec)
sablon_dosya_sec_button.grid(row=1, column=2)

nota8_label = tk.Label(root, text="          ")
nota8_label.grid(row=1, column=3)

# Klasör seçme

kayit_klasoru_label = tk.Label(root, text="\n Toplu Kayıt Klasörü(Word):")
kayit_klasoru_label.grid(row=2, column=0)

kayit_klasoru_entry = tk.Entry(root, width=50)
kayit_klasoru_entry.grid(row=2, column=1)

kayıt_klasor_sec_button = tk.Button(root, text="Klasör Seç", command=klasor_dosyasini_sec)
kayıt_klasor_sec_button.grid(row=2, column=2)

nota7_label = tk.Label(root, text="          ")
nota7_label.grid(row=2, column=3)

nota6_label = tk.Label(root, text=" \n")
nota6_label.grid(row=3, column=0)

toplu_dosya_olustur_button = tk.Button(root, text="Toplu Dosya Oluştur(Word)", command=toplu_word_dosyalari_olustur)
toplu_dosya_olustur_button.grid(row=5, column=0)

toplu_pdfdosya_olustur_button = tk.Button(root, text="WORD to PDF", command=toplu_pdf_dosyalari_olustur)
toplu_pdfdosya_olustur_button.grid(row=5, column=1)

nota13_label = tk.Label(root, text=" \n \n")
nota13_label.grid(row=6, column=0)

toplu_exceldosya_olustur_button = tk.Button(root, text="CSV to Excel", command=csv_to_excel)
toplu_exceldosya_olustur_button.grid(row=6, column=1)

nota3_label = tk.Label(root, text="NOTLAR")
nota3_label.grid(row=9, column=1)
nota3_label.configure(background="white", foreground="black", font=("Arial", 15))

# Resmi yükleyin
image1 = Image.open("resim1.png")

# Boyutları istediğiniz gibi güncelleyin
new_width = 200
new_height = 50
image1 = image1.resize((new_width, new_height))

photo1 = ImageTk.PhotoImage(image1)

# Etiketi oluşturun ve resmi ekleyin
label = tk.Label(root, image=photo1)
label.grid(row=11, column=0)




nota24_label = tk.Label(root, text="\n \n")
nota24_label.grid(row=15, column=1)
root.update()
root.mainloop()
